# -*- coding: utf-8 -*-
import io
from datetime import datetime, time, date
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _excel_safe_datetime(val):
    """Retorna datetime/time/date sem timezone para o Excel."""
    if val is None:
        return None
    if isinstance(val, datetime) and val.tzinfo is not None:
        return val.replace(tzinfo=None)
    if isinstance(val, time) and val.tzinfo is not None:
        return val.replace(tzinfo=None)
    return val


# Usa config.VALOR_DIVISOR (valores no banco em escala 1000x). Ajuste em config/settings.py ou env VALOR_DIVISOR.
from django.conf import settings
VALOR_DIVISOR = getattr(settings, 'VALOR_DIVISOR', 1000)


def _decimal_to_float(value, divisor=1):
    """Converte Decimal para float de forma segura; retorna None em caso de InvalidOperation."""
    if value is None:
        return None
    try:
        if isinstance(value, Decimal):
            if value.is_nan() or value.is_infinite():
                return None
            return float(value) / divisor
        return float(value) / divisor
    except (InvalidOperation, TypeError, ValueError):
        return None


def _safe_cell_value(instance, attr, as_decimal=False, divisor=1):
    """Obtém o valor do atributo do modelo e, se as_decimal, converte para float. Nunca levanta."""
    try:
        val = getattr(instance, attr, None)
        if as_decimal:
            return _decimal_to_float(val, divisor=divisor)
        # Excel não suporta timezone em datetime/time
        val = _excel_safe_datetime(val)
        return val
    except Exception:
        return None if as_decimal else ''


def export_vendas_excel(queryset, max_rows=60000):
    """
    Exporta o queryset de StgVendas para um arquivo xlsx em memória.
    Permitido apenas até 60 mil linhas de dados (além do cabeçalho).
    A view deve bloquear exportação quando o resultado exceder max_rows.
    """
    rows_list = list(queryset[:max_rows])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendas'

    headers = [
        'Nota', 'Data', 'Filial', 'Supervisor', 'Cliente', 'Produto',
        'Qtd', 'Valor Total', 'Peso', 'Devolução', 'Bonificação',
        'Seção', 'Categoria', 'Subcategoria',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, v in enumerate(rows_list, 2):
        try:
            ws.cell(row=row_idx, column=1, value=_safe_cell_value(v, 'numnota'))
            ws.cell(row=row_idx, column=2, value=_safe_cell_value(v, 'data_faturamento'))
            ws.cell(row=row_idx, column=3, value=_safe_cell_value(v, 'codfilial'))
            ws.cell(row=row_idx, column=4, value=_safe_cell_value(v, 'supervisor'))
            ws.cell(row=row_idx, column=5, value=_safe_cell_value(v, 'cliente'))
            ws.cell(row=row_idx, column=6, value=_safe_cell_value(v, 'produto'))
            ws.cell(row=row_idx, column=7, value=_safe_cell_value(v, 'qtd', as_decimal=True))
            ws.cell(row=row_idx, column=8, value=_safe_cell_value(v, 'valortotal', as_decimal=True, divisor=VALOR_DIVISOR))
            ws.cell(row=row_idx, column=9, value=_safe_cell_value(v, 'peso', as_decimal=True))
            ws.cell(row=row_idx, column=10, value=_safe_cell_value(v, 'vldevolucao', as_decimal=True, divisor=VALOR_DIVISOR))
            ws.cell(row=row_idx, column=11, value=_safe_cell_value(v, 'valorbonificado', as_decimal=True, divisor=VALOR_DIVISOR))
            ws.cell(row=row_idx, column=12, value=_safe_cell_value(v, 'secao'))
            ws.cell(row=row_idx, column=13, value=_safe_cell_value(v, 'categoria'))
            ws.cell(row=row_idx, column=14, value=_safe_cell_value(v, 'subcategoria'))
        except Exception:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col, value=None)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
